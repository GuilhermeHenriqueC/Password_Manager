import os
import hashlib
import base64
import getpass

from openpyxl import Workbook, load_workbook
from cryptography.fernet import Fernet

DATA_FILE = "passwords.xlsx"


def derive_fernet_key(passphrase: str) -> bytes:
    """Derive a 32-byte key for Fernet from a passphrase (e.g., 8-word key)."""
    if not passphrase.strip():
        raise ValueError("Passphrase cannot be empty")
    sha256 = hashlib.sha256(passphrase.encode("utf-8")).digest()
    return base64.urlsafe_b64encode(sha256)


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def encrypt_value(key: bytes, value: str) -> str:
    f = Fernet(key)
    return f.encrypt(value.encode("utf-8")).decode("utf-8")


def decrypt_value(key: bytes, token: str) -> str:
    f = Fernet(key)
    return f.decrypt(token.encode("utf-8")).decode("utf-8")


def init_workbook(path: str) -> None:
    if os.path.exists(path):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "passwords"
    ws.append(["name", "enc_username", "enc_password", "username_sha256", "password_sha256"])
    wb.save(path)


def add_entry():
    print("== Add password entry ==")
    name = input("Name (entry label): ").strip()
    if not name:
        print("Name cannot be empty")
        return

    username = input("Username or email: ").strip()
    if not username:
        print("Username/email cannot be empty")
        return

    password = getpass.getpass("Password: ").strip()
    if not password:
        print("Password cannot be empty")
        return

    key_phrase = input("8-word master key (for encryption): ").strip()
    key = derive_fernet_key(key_phrase)

    encrypted_username = encrypt_value(key, username)
    encrypted_password = encrypt_value(key, password)

    username_hash = sha256_text(username)
    password_hash = sha256_text(password)

    wb = load_workbook(DATA_FILE)
    ws = wb.active
    ws.append([name, encrypted_username, encrypted_password, username_hash, password_hash])
    wb.save(DATA_FILE)

    print("Entry saved successfully to", DATA_FILE)


def retrieve_entry():
    print("== Retrieve password entry ==")
    name = input("Enter Name of entry to retrieve: ").strip()
    if not name:
        print("Name cannot be empty")
        return

    wb = load_workbook(DATA_FILE)
    ws = wb.active

    found = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_name, enc_username, enc_password, username_hash, password_hash = row
        if row_name == name:
            found = True
            key_phrase = input("Enter your 8-word master key: ").strip()
            key = derive_fernet_key(key_phrase)
            try:
                username = decrypt_value(key, enc_username)
                password = decrypt_value(key, enc_password)
                print("\n=== Entry found ===")
                print("Name:", name)
                print("Username/Email:", username)
                print("Password:", password)
                print("Username SHA256:", username_hash)
                print("Password SHA256:", password_hash)
            except Exception as e:
                print("Decryption failed: invalid key or data")
            break

    if not found:
        print("Entry not found for Name:", name)


def list_entries():
    print("== List saved entries ==")
    if not os.path.exists(DATA_FILE):
        print("No entries yet")
        return

    wb = load_workbook(DATA_FILE, read_only=True)
    ws = wb.active
    names = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
    if not names:
        print("No entries in the file")
        return

    for idx, entry_name in enumerate(names, start=1):
        print(f"{idx}. {entry_name}")


def main():
    init_workbook(DATA_FILE)

    while True:
        print("\n== Password Manager ==")
        print("1. Add new password entry")
        print("2. Retrieve entry")
        print("3. List entries")
        print("4. Exit")

        choice = input("Choose an option (1-4): ").strip()

        if choice == "1":
            add_entry()
        elif choice == "2":
            retrieve_entry()
        elif choice == "3":
            list_entries()
        elif choice == "4":
            print("Goodbye")
            break
        else:
            print("Invalid choice. Choose 1-4")


if __name__ == "__main__":
    main()
